# -*- coding: utf-8 -*-
import os
import sys
import tempfile
import base64
import streamlit as st

# Add current directory to path so python can find report_generator.py
sys.path.append(os.path.abspath(os.path.dirname(__file__)))
from report_generator import generate_report
from slide_exporter import export_slides_if_needed
from photo_slide_maker import generate_photo_slides

# Base64 image encoding helper for CSS backgrounds
def get_base64_image(file_path):
    if os.path.exists(file_path):
        with open(file_path, "rb") as f:
            data = f.read()
        return base64.b64encode(data).decode()
    return None

def filter_raw_data_by_instructor(input_path, output_path, instructor_name):
    """
    설문 로우데이터 파일에서 강사명에 매핑되는 행만 걸러내어 output_path에 저장.
    강사 관련 열을 찾지 못하거나 매칭되는 행이 없으면 원본 파일을 그대로 리턴.
    """
    import os
    import re
    import openpyxl
    import csv
    
    if not instructor_name or not os.path.exists(input_path):
        return input_path
        
    ext = os.path.splitext(input_path)[1].lower()
    
    # 강사명 정규화 (순수 이름 매칭용)
    clean_target = re.sub(r'\s+', '', instructor_name)
    pure_target = re.sub(r'(교수|강사|선생님|강사님|교수님)', '', clean_target)
    if not pure_target:
        return input_path

    headers = []
    rows = []
    
    try:
        if ext == ".csv":
            data = None
            for enc in ("utf-8-sig", "utf-8", "cp949", "euc-kr"):
                try:
                    with open(input_path, "r", encoding=enc, newline="") as f:
                        data = list(csv.reader(f))
                    break
                except Exception:
                    continue
            if data is None:
                with open(input_path, "r", encoding="utf-8", errors="replace", newline="") as f:
                    data = list(csv.reader(f))
            data = [r for r in data if any((c or "").strip() for c in r)]
            if not data:
                return input_path
            headers = data[0]
            rows = data[1:]
        else:
            wb = openpyxl.load_workbook(input_path, data_only=True)
            ws = wb[wb.sheetnames[0]]
            headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
            for r in range(2, ws.max_row + 1):
                row = [ws.cell(row=r, column=c).value for c in range(1, ws.max_column + 1)]
                if all(v is None or str(v).strip() == "" for v in row):
                    continue
                rows.append(row)
            wb.close()
    except Exception:
        return input_path

    # 강사 관련 열 자동 감지
    instructor_cols = []
    for idx, h in enumerate(headers):
        if h:
            h_str = str(h).strip()
            if any(k in h_str for k in ["강사", "교수", "선생님", "강의"]):
                if not any(k in h_str.lower() for k in ["타임", "timestamp", "time"]):
                    instructor_cols.append(idx)
                    
    if not instructor_cols:
        return input_path

    filtered_rows = []
    for r in rows:
        match = False
        for col_idx in instructor_cols:
            if col_idx < len(r) and r[col_idx] is not None:
                val_str = re.sub(r'\s+', '', str(r[col_idx]))
                if pure_target in val_str:
                    match = True
                    break
        if match:
            filtered_rows.append(r)

    if not filtered_rows:
        return input_path

    try:
        if ext == ".csv":
            with open(output_path, "w", encoding="utf-8-sig", newline="") as f:
                writer = csv.writer(f)
                writer.writerow(headers)
                writer.writerows(filtered_rows)
        else:
            new_wb = openpyxl.Workbook()
            new_ws = new_wb.active
            new_ws.title = "Filtered"
            new_ws.append(headers)
            for r in filtered_rows:
                new_ws.append(r)
            new_wb.save(output_path)
            new_wb.close()
        return output_path
    except Exception:
        return input_path

# Configure Streamlit page layout and aesthetics
st.set_page_config(
    page_title="교육만족도 결과보고서생성기",
    page_icon="📊",
    layout="wide",
    initial_sidebar_state="expanded"
)

# Custom premium styling using CSS
st.markdown(
    """
    <style>
    @import url('https://fonts.googleapis.com/css2?family=Outfit:wght@300;400;600;700&family=Noto+Sans+KR:wght@300;400;700&display=swap');
    
    html, body, [class*="css"] {
        font-family: 'Outfit', 'Noto Sans KR', sans-serif;
    }
    
    .main-title {
        font-size: 2.8rem;
        font-weight: 700;
        background: linear-gradient(135deg, #2D6CDF, #00C9FF);
        -webkit-background-clip: text;
        -webkit-text-fill-color: transparent;
        margin-bottom: 0.5rem;
    }
    
    .subtitle {
        font-size: 1.1rem;
        color: #6c757d;
        margin-bottom: 2rem;
    }
    
    .card {
        background-color: #ffffff;
        border-radius: 12px;
        padding: 24px;
        box-shadow: 0 4px 6px rgba(0, 0, 0, 0.05);
        border: 1px solid #eef1f6;
        margin-bottom: 20px;
    }
    
    .stButton>button {
        background: linear-gradient(135deg, #2D6CDF, #1B4FB0);
        color: white;
        border-radius: 8px;
        padding: 10px 24px;
        font-weight: 600;
        border: none;
        box-shadow: 0 4px 10px rgba(45, 108, 223, 0.2);
        transition: all 0.3s ease;
    }
    
    .stButton>button:hover {
        transform: translateY(-2px);
        box-shadow: 0 6px 15px rgba(45, 108, 223, 0.3);
    }
    
    .sidebar-header {
        font-size: 1.3rem;
        font-weight: 700;
        color: #2D6CDF;
        margin-bottom: 15px;
    }
    
    .sidebar-section {
        background-color: #f8f9fa;
        padding: 15px;
        border-radius: 8px;
        margin-bottom: 15px;
        border-left: 4px solid #2D6CDF;
    }
    
    .download-link-btn {
        display: inline-block;
        width: 100%;
        text-align: center;
        background-color: #f1f3f7;
        color: #2D6CDF;
        padding: 8px 12px;
        border-radius: 6px;
        text-decoration: none;
        font-weight: 600;
        margin-bottom: 8px;
        border: 1px solid #DFE4EC;
        transition: all 0.2s ease;
    }

    .download-link-btn:hover {
        background-color: #EAF1FF;
        border-color: #2D6CDF;
    }

    /* 가상 PPT 슬라이드 프리뷰 스타일 */
    .ppt-preview-container {
        width: 100%;
        background-color: #f1f3f7;
        padding: 12px;
        border-radius: 12px;
        box-shadow: inset 0 2px 5px rgba(0,0,0,0.05);
        display: flex;
        justify-content: center;
        align-items: center;
        margin-top: 10px;
        margin-bottom: 20px;
    }
    .ppt-editor-frame {
        width: 100%;
        border: 2px solid #D2D6DC;
        border-radius: 10px;
        background-color: #E5E9F0;
        padding: 6px;
        box-shadow: 0 8px 20px rgba(0,0,0,0.06);
    }
    .ppt-editor-header-bar {
        display: flex;
        align-items: center;
        justify-content: space-between;
        background-color: #ffffff;
        padding: 6px 12px;
        border-radius: 6px 6px 0 0;
        border-bottom: 1.5px solid #E5E9F0;
        font-size: 0.75rem;
        font-weight: 600;
        color: #4A5568;
    }
    .ppt-editor-dots {
        display: flex;
        gap: 5px;
    }
    .ppt-dot {
        width: 8px;
        height: 8px;
        border-radius: 50%;
        background-color: #CBD5E0;
    }
    .ppt-dot.red { background-color: #FC8181; }
    .ppt-dot.yellow { background-color: #F6E05E; }
    .ppt-dot.green { background-color: #68D391; }
    
    .ppt-preview-slide {
        width: 100%;
        background-color: #ffffff;
        border: 1px solid #DFE4EC;
        border-radius: 0 0 6px 6px;
        padding: 28px 20px 20px 20px;
        position: relative;
        box-sizing: border-box;
        display: flex;
        flex-direction: column;
        justify-content: space-between;
        min-height: 260px;
    }
    .ppt-slide-badge {
        position: absolute;
        top: 10px;
        right: 10px;
        background: linear-gradient(135deg, #FF6B6B, #FF8E53);
        color: white;
        font-size: 0.65rem;
        font-weight: 700;
        padding: 2px 7px;
        border-radius: 20px;
        box-shadow: 0 2px 5px rgba(255, 107, 107, 0.25);
        z-index: 10;
        line-height: 1.2;
    }
    .ppt-preview-blue-box {
        border: 1.5px dashed #2D6CDF;
        background-color: rgba(45, 108, 223, 0.03);
        border-radius: 6px;
        padding: 8px 12px;
        position: relative;
        margin-bottom: 12px;
        transition: all 0.2s ease;
    }
    .ppt-preview-blue-box:hover {
        background-color: rgba(45, 108, 223, 0.08);
        border-style: solid;
        box-shadow: 0 0 8px rgba(45, 108, 223, 0.15);
    }
    .box-tag {
        position: absolute;
        top: -8px;
        left: 8px;
        background-color: #2D6CDF;
        color: white;
        font-size: 0.6rem;
        font-weight: 700;
        padding: 1px 6px;
        border-radius: 3px;
        line-height: 1.2;
    }
    .box-content {
        font-size: 0.85rem;
        color: #333333;
        font-weight: 500;
        min-height: 14px;
        word-break: break-all;
    }
    .box-content-title1 {
        font-size: 0.95rem;
        font-weight: 400;
        color: #555555;
        margin-bottom: 2px;
        min-height: 16px;
        word-break: break-all;
    }
    .box-content-title2 {
        font-size: 1.25rem;
        font-weight: 700;
        color: #111111;
        min-height: 24px;
        word-break: break-all;
    }
    .ppt-preview-footer {
        display: grid;
        grid-template-columns: 1.2fr 0.8fr;
        gap: 10px;
        margin-top: auto;
    }
    .overview-slide {
        justify-content: flex-start;
    }
    .slide-header {
        font-size: 0.95rem;
        font-weight: 700;
        color: #2D6CDF;
        margin-bottom: 12px;
        border-bottom: 2px solid #2D6CDF;
        padding-bottom: 4px;
    }
    .ppt-preview-table {
        width: 100%;
        border-collapse: collapse;
        font-size: 0.8rem;
    }
    .ppt-preview-table tr {
        border-bottom: 1px solid #f1f3f7;
    }
    .table-label {
        width: 25%;
        font-weight: 700;
        color: #4a5568;
        padding: 6px 0;
        vertical-align: middle;
    }
    .table-value {
        width: 75%;
        padding: 4px 0;
    }
    .inline-box {
        margin-bottom: 0;
        padding: 6px 10px;
    }

    /* 실제 슬라이드 오버레이용 CSS */
    .ppt-slide-overlay-container {
        position: relative;
        width: 100%;
        aspect-ratio: 16 / 9;
        background-color: #ffffff;
        background-size: 100% 100%;
        background-repeat: no-repeat;
        border: 1px solid #DFE4EC;
        border-radius: 0 0 6px 6px;
        box-sizing: border-box;
        overflow: hidden;
        min-height: 250px;
    }
    
    /* 절대 좌표 블루 박스 */
    .ppt-abs-box {
        position: absolute;
        border: 1.8px dashed #E53E3E;
        background-color: rgba(229, 62, 62, 0.06);
        border-radius: 4px;
        padding: 4px 6px;
        box-sizing: border-box;
        transition: all 0.2s ease;
        z-index: 5;
        overflow: hidden;
    }
    .ppt-abs-box:hover {
        background-color: rgba(229, 62, 62, 0.12);
        border-style: solid;
        box-shadow: 0 0 8px rgba(229, 62, 62, 0.3);
    }
    
    /* 표지 절대 좌표 배치 */
    .pin-cover-label {
        top: 6%;
        left: 7%;
        width: 25%;
        height: 12%;
    }
    .pin-cover-title {
        top: 26%;
        left: 7%;
        width: 86%;
        height: 32%;
    }
    .pin-cover-instructor {
        top: 73%;
        left: 7%;
        width: 44%;
        height: 14%;
    }
    .pin-cover-students {
        top: 73%;
        left: 54%;
        width: 39%;
        height: 14%;
    }
    
    /* 교육개요 절대 좌표 배치 */
    .pin-overview-name {
        top: 23%;
        left: 27%;
        width: 66%;
        height: 9%;
    }
    .pin-overview-schedule {
        top: 34%;
        left: 27%;
        width: 66%;
        height: 9%;
    }
    .pin-overview-method {
        top: 45%;
        left: 27%;
        width: 66%;
        height: 9%;
    }
    .pin-overview-target {
        top: 56%;
        left: 27%;
        width: 66%;
        height: 9%;
    }
    .pin-overview-goal {
        top: 67%;
        left: 27%;
        width: 66%;
        height: 22%;
    }
    </style>
    """,
    unsafe_allow_html=True
)


def get_file_bytes(filepath):
    if os.path.exists(filepath):
        with open(filepath, "rb") as f:
            return f.read()
    return None


# Locate default files
current_dir = os.path.dirname(os.path.abspath(__file__))

default_template_path = os.path.join(current_dir, "템플릿_결보표양.pptx")
if not os.path.exists(default_template_path):
    default_template_path = os.path.join(current_dir, "assets", "템플릿_결보표양.pptx")

default_config_path = os.path.join(current_dir, "설정.xlsx")
if not os.path.exists(default_config_path):
    default_config_path = os.path.join(current_dir, "assets", "설정_예시.xlsx")

# PPT 슬라이드 이미지 사전 추출 구동 (캐싱 내장)
export_slides_if_needed(default_template_path, os.path.join(current_dir, "assets"))

# Base64 이미지 변수화
slide1_base64 = get_base64_image(os.path.join(current_dir, "assets", "slide1_origin.png"))
slide4_base64 = get_base64_image(os.path.join(current_dir, "assets", "slide4_origin.png"))



# ---------------- Sidebar ----------------
with st.sidebar:
    st.markdown("<div class='sidebar-header'>📊 KOOS RESULT</div>", unsafe_allow_html=True)
    st.markdown(
        "<p style='font-size: 0.9rem; color:#666;'>교육 만족도 결과보고서 PPT를 자동생성해주는 웹 서비스입니다.</p>",
        unsafe_allow_html=True
    )

    st.markdown("---")
    menu = st.selectbox(
        "🛠️ 기능 선택",
        ["📊 결과보고서 자동 생성", "📸 사진 슬라이드 생성 (PhotoSlideMaker)"],
        index=0,
        help="만족도 조사 결과를 기반으로 결과보고서 PPT를 자동 빌드하거나, 교육 활동 사진을 업로드해 사진 스케치 슬라이드를 만듭니다."
    )

    st.markdown("---")
    st.markdown("<div style='font-weight:600; margin-bottom:10px;'>리소스 다운로드</div>", unsafe_allow_html=True)

    template_bytes = get_file_bytes(default_template_path)
    if template_bytes:
        st.download_button(
            label="📄 기본 PPT 템플릿 다운로드",
            data=template_bytes,
            file_name="템플릿_결보표양.pptx",
            mime="application/vnd.openxmlformats-officedocument.presentationml.presentation",
            use_container_width=True
        )

    config_bytes = get_file_bytes(default_config_path)
    if config_bytes:
        st.download_button(
            label="⚙️ 기본 설정 엑셀 다운로드",
            data=config_bytes,
            file_name="설정.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True
        )

    st.markdown("---")
    st.markdown("<div style='font-weight:600; margin-bottom:10px;'>이용 안내</div>", unsafe_allow_html=True)

    with st.expander("💡 척도 자동 감지 규칙"):
        st.markdown(
            """
            - 5점 그렇다형, 만족형, 우수형, 도움형
            - 4점 그렇다형, 만족형
            - 7점 그렇다형 등을 자동 감지합니다.
            - 응답의 60% 이상이 특정 척도 라벨을 가질 때 작동합니다.
            """
        )

    with st.expander("💡 주관식 자동 분류"):
        st.markdown(
            """
            - 문항 제목에 `좋았던·장점·우수·인상` 등이 포함되면 **좋았던 점**으로 분류됩니다.
            - `아쉬·개선·보완·불편·단점·건의` 등이 포함되면 **아쉬웠던 점**으로 분류됩니다.
            - 아쉬웠던 점의 '없음', '없습니다' 등 무의미한 답변은 기본적으로 자동 제외됩니다.
            """
        )


# ---------------- Main Page ----------------
if menu == "📸 사진 슬라이드 생성 (PhotoSlideMaker)":
    st.markdown("<div class='main-title'>📸 교육 사진 슬라이드 자동 생성기 (PhotoSlideMaker)</div>", unsafe_allow_html=True)
    st.markdown(
        "<div class='subtitle'>교육 현장 사진들을 업로드하면, 장수에 알맞은 최적의 격자 그리드로 배치된 스케치 슬라이드 PPTX가 즉시 빌드됩니다.</div>",
        unsafe_allow_html=True
    )

    col_ph_left, col_ph_right = st.columns([1.1, 0.9])

    with col_ph_left:
        st.markdown("### 1. 사진 리소스 업로드")
        photo_files = st.file_uploader(
            "📁 사진 파일들(다중 선택 가능) 또는 ZIP 압축 파일 업로드 (필수)",
            type=["zip", "jpg", "jpeg", "png", "webp", "bmp"],
            accept_multiple_files=True,
            help="여러 장의 이미지 파일을 다중 선택하여 올리거나, 사진 폴더를 ZIP으로 압축해 올릴 수 있습니다.",
            key="photo_files_uploader"
        )

        st.markdown("### 2. 슬라이드 설정")
        max_photos = st.slider(
            "🎛️ 슬라이드당 최대 사진 수",
            min_value=1,
            max_value=6,
            value=6,
            help="한 페이지에 최대 몇 장의 사진을 격자 배치할지 선택합니다. 6장을 초과하면 다음 슬라이드에 자동 분할 배치됩니다."
        )

        use_photo_template = st.checkbox(
            "🎨 기존 PPT 템플릿(PhotoSlides.pptx) 뒤에 이어 붙이기",
            value=False,
            help="체크할 경우 업로드하신 PPT 파일 맨 뒤에 사진 슬라이드를 덧붙여서 새로 생성합니다."
        )

        photo_template_file = None
        if use_photo_template:
            photo_template_file = st.file_uploader(
                "🎨 템플릿 PPTX 파일 업로드",
                type=["pptx"],
                help="사진 슬라이드를 뒤에 추가해 붙여넣을 대상 마스터 파워포인트 문서입니다.",
                key="photo_template_uploader"
            )

    with col_ph_right:
        st.markdown("### 3. 슬라이드 빌드 실행")
        build_btn = st.button("사진 슬라이드 PPTX 생성하기 🚀", use_container_width=True, key="photo_build_btn")

        photo_log_area = st.empty()
        photo_dl_area = st.empty()

        if build_btn:
            if not photo_files:
                st.error("❌ 업로드된 사진 파일이나 ZIP 파일이 없습니다.")
            else:
                with tempfile.TemporaryDirectory() as tmpdir:
                    photo_logs = []
                    def photo_log(msg):
                        photo_logs.append(msg)
                        photo_log_area.code("\n".join(photo_logs[-10:]))

                    # 단일 ZIP 파일 처리
                    if len(photo_files) == 1 and photo_files[0].name.lower().endswith(".zip"):
                        target_input_path = os.path.join(tmpdir, photo_files[0].name)
                        with open(target_input_path, "wb") as f:
                            f.write(photo_files[0].getbuffer())
                    else:
                        # 개별 사진 다중 선택 처리
                        target_input_path = os.path.join(tmpdir, "images")
                        os.makedirs(target_input_path, exist_ok=True)
                        for uploaded_f in photo_files:
                            with open(os.path.join(target_input_path, uploaded_f.name), "wb") as f:
                                f.write(uploaded_f.getbuffer())

                    # 템플릿 처리
                    tpl_path = None
                    if use_photo_template and photo_template_file:
                        tpl_path = os.path.join(tmpdir, "photo_template.pptx")
                        with open(tpl_path, "wb") as f:
                            f.write(photo_template_file.getbuffer())

                    output_ppt_path = os.path.join(tmpdir, "PhotoSlides_output.pptx")

                    try:
                        with st.spinner("사진 자동 배치 및 슬라이드 빌드 중..."):
                            generate_photo_slides(
                                src_path=target_input_path,
                                template_path=tpl_path,
                                output_path=output_ppt_path,
                                max_per_slide=max_photos,
                                log=photo_log
                            )

                        if os.path.exists(output_ppt_path):
                            with open(output_ppt_path, "rb") as f:
                                result_bytes = f.read()

                            st.success("🎉 사진 슬라이드 PPTX 생성이 완료되었습니다!")
                            photo_dl_area.download_button(
                                label="📥 사진 슬라이드 다운로드 (.pptx)",
                                data=result_bytes,
                                file_name="교육사진_스케치_슬라이드.pptx",
                                mime="application/vnd.openxmlformats-officedocument.presentationml.presentation",
                                use_container_width=True
                            )
                        else:
                            st.error("❌ 사진 슬라이드 파일 생성에 실패했습니다.")
                    except Exception as ex:
                        st.error(f"❌ 빌드 도중 에러 발생: {str(ex)}")

    st.stop()

st.markdown("<div class='main-title'>교육만족도 결과보고서생성기</div>", unsafe_allow_html=True)
st.markdown(
    "<div class='subtitle'>설문 로우데이터와 설정 값만 입력하면 시각화 결과 보고서 PPT가 즉시 만들어집니다.</div>",
    unsafe_allow_html=True
)

col_left, col_right = st.columns([1.1, 0.9])

with col_left:
    st.markdown("### 1. 입력 파일 업로드")

    config_mode_val = st.session_state.get("config_mode", "직접 입력 (추천 - 엑셀 파일 없이 직접 타이핑)")
    instructor_count = st.session_state.get("instructor_count", 1)
    
    raw_files = {}
    if "직접 입력" in config_mode_val and instructor_count > 1:
        st.markdown("##### 📊 강사별 설문 로우데이터 업로드")
        for i in range(instructor_count):
            inst_name = st.session_state.get(f"instructor_{i}", "설상훈 교수" if i == 0 else "").strip()
            if not inst_name:
                inst_name = f"강사 {i+1}"
            raw_files[inst_name] = st.file_uploader(
                f"📂 [{inst_name}] 설문 로우데이터 업로드 (xlsx, csv)",
                type=["xlsx", "xls", "xlsm", "csv"],
                key=f"raw_file_{i}"
            )
    else:
        raw_file = st.file_uploader(
            "📊 설문 로우데이터 엑셀 또는 CSV 파일 업로드 (필수)",
            type=["xlsx", "xls", "xlsm", "csv"],
            help="구글폼 등에서 내려받은 응답 원본 파일입니다."
        )
        if raw_file:
            inst_name = st.session_state.get("instructor_0", "설상훈 교수").strip()
            if not inst_name:
                inst_name = "강사"
            raw_files[inst_name] = raw_file

    valid_raw_files = {name: f for name, f in raw_files.items() if f is not None}

    if valid_raw_files:
        st.markdown('<div class="card" style="padding: 15px; background-color: #EAF1FF; border: 1px solid #2D6CDF; margin-top: 10px; margin-bottom: 15px;">', unsafe_allow_html=True)
        col_lbl, col_btn = st.columns([1.2, 0.8])
        with col_lbl:
            st.markdown("<p style='margin:0; font-weight:600; color:#2D6CDF; font-size:0.9rem;'>💬 업로드된 파일의 주관식 답변에서 '없음' 의견 비중을 사전 분석해보세요.</p>", unsafe_allow_html=True)
        with col_btn:
            analyze_subj = st.button("🔍 주관식 '없음' 통계 분석", use_container_width=True)
        st.markdown('</div>', unsafe_allow_html=True)

        if analyze_subj:
            with st.spinner("주관식 의견 분석 중..."):
                try:
                    with tempfile.TemporaryDirectory() as tmpdir:
                        inst_names = list(valid_raw_files.keys())
                        tabs = st.tabs([f"🏫 {name}" for name in inst_names])
                        
                        for idx, inst_name in enumerate(inst_names):
                            uploaded_file = valid_raw_files[inst_name]
                            raw_ext = os.path.splitext(uploaded_file.name)[1]
                            raw_temp_path = os.path.join(tmpdir, f"raw_data_{idx}{raw_ext}")
                            with open(raw_temp_path, "wb") as f:
                                f.write(uploaded_file.getbuffer())

                            # 통합 로우데이터 대응: 강사명 필터링
                            filtered_temp_path = os.path.join(tmpdir, f"filtered_data_{idx}{raw_ext}")
                            actual_data_path = filter_raw_data_by_instructor(raw_temp_path, filtered_temp_path, inst_name)

                            from report_generator import parse_raw_data, _is_no_opinion
                            data = parse_raw_data(actual_data_path)
                            subjective = data.get("subjective", [])
                            n_resp = data.get("n_responses", 0)

                            with tabs[idx]:
                                if not subjective:
                                    st.warning(f"⚠️ {inst_name} 파일 내에서 분석할 수 있는 주관식 문항을 찾지 못했습니다.")
                                else:
                                    st.markdown(f"##### 📊 {inst_name} 주관식 통계 분석 결과 (총 응답자: {n_resp}명)")
                                    
                                    for sq in subjective:
                                        total_ans = len(sq["answers"])
                                        no_op_count = sum(1 for ans in sq["answers"] if _is_no_opinion(ans))
                                        no_op_rate = round(no_op_count / total_ans * 100) if total_ans else 0
                                        
                                        # "없" 단어 포함 의견 분석
                                        no_word_answers = [ans for ans in sq["answers"] if "없" in ans]
                                        no_word_count = len(no_word_answers)
                                        no_word_rate = round(no_word_count / total_ans * 100) if total_ans else 0
                                        
                                        st.markdown(f"""
                                        <div style="background-color: #ffffff; border-radius: 8px; border: 1px solid #eef1f6; box-shadow: 0 4px 6px rgba(0,0,0,0.02); padding: 16px; margin-bottom: 20px; border-top: 4px solid #2D6CDF;">
                                            <div style="font-weight: 700; font-size: 1rem; color: #1A202C; margin-bottom: 12px; line-height: 1.4;">{sq['header']}</div>
                                            <div style="display: grid; grid-template-columns: 1fr 1fr; gap: 20px;">
                                                <div style="font-size: 0.85rem; color: #4A5568; display: flex; flex-direction: column; gap: 8px;">
                                                    <div>💬 전체 응답: <strong>{total_ans}건</strong></div>
                                                    <div style="color: #E53E3E;">🚫 '없음'류(무의견): <strong>{no_op_count}건 ({no_op_rate}%)</strong></div>
                                                    <div style="color: #2D6CDF;">✅ 유효 의견: <strong>{total_ans - no_op_count}건 ({100 - no_op_rate}%)</strong></div>
                                                    <hr style="margin: 8px 0; border: none; border-top: 1px solid #eef1f6;" />
                                                    <div style="color: #DD6B20; font-weight: 600;">🔍 "없" 글자 포함 의견: {no_word_count}건 ({no_word_rate}%)</div>
                                                    <div style="background-color: #EDF2F7; border-radius: 4px; width: 100%; height: 8px; overflow: hidden; margin-top: 4px;">
                                                        <div style="background-color: #DD6B20; width: {no_word_rate}%; height: 100%;"></div>
                                                    </div>
                                                </div>
                                                <div style="border-left: 1px solid #E2E8F0; padding-left: 20px;">
                                                    <div style="font-size: 0.85rem; font-weight: 600; color: #2D3748; margin-bottom: 6px;">📋 "없" 포함 의견 모아보기</div>
                                                    <div style="overflow-y: auto; max-height: 140px; background-color: #F7FAFC; border: 1px solid #EDF2F7; border-radius: 6px; padding: 10px; font-size: 0.8rem; line-height: 1.4;">
                                                        {"".join([f'<div style="margin-bottom: 6px; color: #4A5568; border-bottom: 1px dashed #EDF2F7; padding-bottom: 4px;">• {ans}</div>' for ans in no_word_answers]) if no_word_answers else '<div style="color: #A0AEC0; text-align: center; padding-top: 10px;">"없" 단어가 포함된 의견이 없습니다.</div>'}
                                                    </div>
                                                </div>
                                            </div>
                                        </div>
                                        """, unsafe_allow_html=True)
                except Exception as e:
                    st.error(f"❌ 분석 중 오류 발생: {str(e)}")

    st.markdown("---")
    st.markdown("### 2. 보고서 세부 정보 설정")

    config_mode = st.radio(
        "설정 값 입력 방식 선택",
        options=["직접 입력 (추천 - 엑셀 파일 없이 직접 타이핑)", "엑셀 파일 업로드 (설정.xlsx 사용)"],
        horizontal=True,
        key="config_mode"
    )

    config_dict = None
    config_file = None

    if "직접 입력" in config_mode:
        st.info("💡 아래 입력 칸에 정보를 적어주세요. 우측 가상 슬라이드(실제 PPT 매핑 가이드)에 위치가 실시간 시각화됩니다.")

        tab_basic, tab_overview = st.tabs(["📝 표지 슬라이드 매핑 가이드", "🏫 교육개요 슬라이드 매핑 가이드"])

        with tab_basic:
            col_in, col_pre = st.columns([1.1, 0.9])
            
            with col_in:
                cover_label = st.text_input(
                    "표지 상단 라벨", 
                    value="결과보고서",
                    help="📄 실제 PPT 파일 [1페이지(표지)] 맨 위에 위치한 회색 박스 영역에 매핑됩니다."
                )
                title1 = st.text_input(
                    "표지 대제목 1줄", 
                    value="SK TNS FLP 대상",
                    help="📄 실제 PPT 파일 [1페이지(표지)] 중앙 대제목의 첫 번째 줄에 매핑됩니다."
                )
                title2 = st.text_input(
                    "표지 대제목 2줄", 
                    value="인사이트 트립 결과보고서",
                    help="📄 실제 PPT 파일 [1페이지(표지)] 중앙 대제목의 두 번째 줄에 매핑됩니다."
                )

                if "instructor_count" not in st.session_state:
                    st.session_state.instructor_count = 1

                col_inst_title, col_inst_plus, col_inst_minus = st.columns([5, 1, 1])

                with col_inst_title:
                    st.markdown("**강사명** (표지 우하단 반영)")

                with col_inst_plus:
                    add_disabled = st.session_state.instructor_count >= 5
                    if st.button("➕", help="강사 추가", disabled=add_disabled, key="add_inst"):
                        st.session_state.instructor_count += 1
                        st.rerun()

                with col_inst_minus:
                    minus_disabled = st.session_state.instructor_count <= 1
                    if st.button("➖", help="강사 삭제", disabled=minus_disabled, key="del_inst"):
                        st.session_state.instructor_count -= 1
                        st.rerun()

                instructors = []
                for i in range(st.session_state.instructor_count):
                    default_value = "설상훈 교수" if i == 0 else ""
                    instructor_name = st.text_input(
                        f"강사 {i + 1}",
                        value=default_value,
                        key=f"instructor_{i}",
                        help="📄 실제 PPT 파일 [1페이지(표지)] 우하단 강사명 영역에 매핑됩니다."
                    )
                    instructors.append(instructor_name)

                instructor = ", ".join([name.strip() for name in instructors if name and name.strip()])

                student_count = st.number_input(
                    "수강인원(명)",
                    min_value=1,
                    value=9,
                    step=1,
                    help="📄 실제 PPT 파일 [1페이지(표지)] 우하단 인원 수에 매핑되며, [7페이지]의 응답률(수강인원 대비 응답자 수) 계산에도 자동 대입됩니다."
                )

            with col_pre:
                st.markdown("**🎨 PPT 표지 슬라이드(1페이지) 실제 매핑 맵**")
                bg_style = f"background-image: url(data:image/png;base64,{slide1_base64});" if slide1_base64 else "background-color: #ffffff;"
                preview_html = f"""
                <div class="ppt-preview-container">
                    <div class="ppt-editor-frame">
                        <div class="ppt-editor-header-bar">
                            <div class="ppt-editor-dots">
                                <span class="ppt-dot red"></span>
                                <span class="ppt-dot yellow"></span>
                                <span class="ppt-dot green"></span>
                            </div>
                            <div style="font-size: 0.7rem; color: #718096; font-family: monospace;">템플릿_결보표양.pptx - 슬라이드 1 (표지)</div>
                            <div></div>
                        </div>
                        <div class="ppt-slide-overlay-container" style="{bg_style}">
                            <span class="ppt-slide-badge">슬라이드 1 / 9</span>
                            <div class="ppt-abs-box pin-cover-label">
                                <span class="box-tag" style="background-color: #E53E3E;">표지 상단 라벨 (1p 맨위)</span>
                                <div class="box-content">{cover_label if cover_label else '&nbsp;'}</div>
                            </div>
                            <div class="ppt-abs-box pin-cover-title">
                                <span class="box-tag" style="background-color: #E53E3E;">표지 대제목 1~2줄 (1p 중앙)</span>
                                <div class="box-content-title1">{title1 if title1 else '&nbsp;'}</div>
                                <div class="box-content-title2">{title2 if title2 else '&nbsp;'}</div>
                            </div>
                            <div class="ppt-abs-box pin-cover-instructor">
                                <span class="box-tag" style="background-color: #E53E3E;">강사명 (1p 우하단)</span>
                                <div class="box-content">{instructor if instructor else '&nbsp;'}</div>
                            </div>
                            <div class="ppt-abs-box pin-cover-students">
                                <span class="box-tag" style="background-color: #E53E3E;">수강인원 (1p 우하단)</span>
                                <div class="box-content">{student_count}명</div>
                            </div>
                        </div>
                    </div>
                </div>
                """
                st.markdown(preview_html, unsafe_allow_html=True)

        with tab_overview:
            col_in, col_pre = st.columns([1.1, 0.9])
            
            with col_in:
                course_name = st.text_input(
                    "과정명", 
                    value="인사이트 트립",
                    help="📄 실제 PPT 파일 [4페이지(교육개요 표)]의 '과정명' 행의 두 번째 열에 대입됩니다."
                )
                schedule = st.text_input(
                    "교육일정", 
                    value="2026.07.08",
                    help="📄 실제 PPT 파일 [4페이지(교육개요 표)]의 '교육일정' 행의 두 번째 열에 대입됩니다."
                )
                method = st.text_input(
                    "교육방식", 
                    value="대면 교육",
                    help="📄 실제 PPT 파일 [4페이지(교육개요 표)]의 '교육방식' 행의 두 번째 열에 대입됩니다."
                )
                target = st.text_input(
                    "교육대상", 
                    value="핵심 실무진",
                    help="📄 실제 PPT 파일 [4페이지(교육개요 표)]의 '교육 대상' 행의 두 번째 열에 대입됩니다."
                )
                goal = st.text_area(
                    "교육목표", 
                    value="트렌드 학습 및 비즈니스 인사이트 도출",
                    help="📄 실제 PPT 파일 [4페이지(교육개요 표)]의 '교육 목표' 행의 두 번째 열에 대입됩니다."
                )

            with col_pre:
                st.markdown("**🎨 PPT 교육개요 슬라이드(4페이지) 실제 매핑 맵**")
                bg_style = f"background-image: url(data:image/png;base64,{slide4_base64});" if slide4_base64 else "background-color: #ffffff;"
                preview_html = f"""
                <div class="ppt-preview-container">
                    <div class="ppt-editor-frame">
                        <div class="ppt-editor-header-bar">
                            <div class="ppt-editor-dots">
                                <span class="ppt-dot red"></span>
                                <span class="ppt-dot yellow"></span>
                                <span class="ppt-dot green"></span>
                            </div>
                            <div style="font-size: 0.7rem; color: #718096; font-family: monospace;">템플릿_결보표양.pptx - 슬라이드 4 (교육개요)</div>
                            <div></div>
                        </div>
                        <div class="ppt-slide-overlay-container" style="{bg_style}">
                            <span class="ppt-slide-badge">슬라이드 4 / 9</span>
                            <div class="ppt-abs-box pin-overview-name">
                                <span class="box-tag" style="background-color: #E53E3E;">과정명 (4p 표 1행)</span>
                                <div class="box-content">{course_name if course_name else '&nbsp;'}</div>
                            </div>
                            <div class="ppt-abs-box pin-overview-schedule">
                                <span class="box-tag" style="background-color: #E53E3E;">교육일정 (4p 표 2행)</span>
                                <div class="box-content">{schedule if schedule else '&nbsp;'}</div>
                            </div>
                            <div class="ppt-abs-box pin-overview-method">
                                <span class="box-tag" style="background-color: #E53E3E;">교육방식 (4p 표 3행)</span>
                                <div class="box-content">{method if method else '&nbsp;'}</div>
                            </div>
                            <div class="ppt-abs-box pin-overview-target">
                                <span class="box-tag" style="background-color: #E53E3E;">교육대상 (4p 표 4행)</span>
                                <div class="box-content">{target if target else '&nbsp;'}</div>
                            </div>
                            <div class="ppt-abs-box pin-overview-goal">
                                <span class="box-tag" style="background-color: #E53E3E;">교육목표 (4p 표 5행)</span>
                                <div class="box-content" style="white-space: pre-wrap;">{goal if goal else '&nbsp;'}</div>
                            </div>
                        </div>
                    </div>
                </div>
                """
                st.markdown(preview_html, unsafe_allow_html=True)

        config_dict = {
            "basic": {
                "표지_상단라벨": cover_label,
                "표지_제목1": title1,
                "표지_제목2": title2,
                "강사명": instructor,
                "수강인원": int(student_count)
            },
            "overview": {
                "과정명": course_name,
                "교육일정": schedule,
                "교육방식": method,
                "교육대상": target,
                "교육목표": goal
            },
            "questions": []
        }

    else:
        st.info("📂 '설정.xlsx' 파일을 업로드해 주세요. (다운로드 탭에서 예시를 내려받아 수정할 수 있습니다.)")
        config_file = st.file_uploader(
            "⚙️ 설정 엑셀 파일 업로드",
            type=["xlsx", "xls", "xlsm"],
            help="표지 제목, 교육 개요 등이 기록된 설정 엑셀 파일입니다."
        )


with col_right:
    st.markdown("### 3. 추가 옵션")

    keep_no_opinion = st.checkbox(
        "아쉬웠던 점의 '없음'류 의견도 그대로 표시",
        value=False,
        help="체크 해제 시 '없음', '없습니다', '해당사항 없음' 같은 짧고 무의미한 의견을 필터링합니다."
    )

    use_custom_template = st.checkbox(
        "사용자 정의 PPT 템플릿 파일 업로드",
        value=False,
        help="체크 시 나만의 PPT 템플릿(.pptx)을 업로드하여 반영할 수 있습니다."
    )

    template_file = None
    if use_custom_template:
        template_file = st.file_uploader(
            "🎨 템플릿 PPTX 파일 업로드",
            type=["pptx"],
            help="결과보고서 마스터 레이아웃이 적용된 파워포인트 템플릿입니다."
        )
    else:
        st.caption(
            f"💡 기본 템플릿 (`{os.path.basename(default_template_path or '템플릿_결보표양.pptx')}`)을 사용합니다."
        )

    st.markdown("---")
    st.markdown("### 4. 보고서 생성")

    generate_btn = st.button("결과보고서 만들기 🚀", use_container_width=True)

    log_area = st.empty()
    download_area = st.empty()

    if generate_btn:
        missing_files = []
        if "직접 입력" in config_mode_val:
            for i in range(instructor_count):
                inst_name = st.session_state.get(f"instructor_{i}", "설상훈 교수" if i == 0 else "").strip()
                if not inst_name:
                    inst_name = f"강사 {i+1}"
                if raw_files.get(inst_name) is None:
                    missing_files.append(inst_name)
        else:
            if not valid_raw_files:
                missing_files.append("설문 로우데이터")

        if missing_files:
            st.error(f"❌ 다음 항목의 로우데이터 파일을 업로드해 주세요: {', '.join(missing_files)}")
        elif "엑셀 파일 업로드" in config_mode_val and not config_file:
            st.error("❌ 설정 엑셀 파일(설정.xlsx)을 업로드해 주세요.")
        elif use_custom_template and not template_file:
            st.error("❌ 템플릿 PPTX 파일을 업로드해 주세요.")
        else:
            with tempfile.TemporaryDirectory() as tmpdir:
                template_temp_path = default_template_path
                if template_file:
                    template_temp_path = os.path.join(tmpdir, "template.pptx")
                    with open(template_temp_path, "wb") as f:
                        f.write(template_file.getbuffer())

                if not template_temp_path or not os.path.exists(template_temp_path):
                    st.error("❌ 기본 템플릿 파일을 찾을 수 없습니다. 템플릿 파일을 직접 업로드해 주세요.")
                else:
                    logs = []

                    def streamlit_log(message):
                        logs.append(message)
                        log_area.code("\n".join(logs[-10:]))

                    pptx_results = {}

                    try:
                        streamlit_log("[시작] 교육보고서 생성을 시작합니다...")

                        # 직접 입력 모드에서 강사별 개별 보고서 순차 빌드
                        if "직접 입력" in config_mode_val:
                            import copy
                            for i in range(instructor_count):
                                inst_name = st.session_state.get(f"instructor_{i}", "설상훈 교수" if i == 0 else "").strip()
                                if not inst_name:
                                    inst_name = f"강사 {i+1}"

                                uploaded_raw = raw_files[inst_name]
                                raw_ext = os.path.splitext(uploaded_raw.name)[1]
                                raw_temp_path = os.path.join(tmpdir, f"raw_{i}{raw_ext}")

                                with open(raw_temp_path, "wb") as f:
                                    f.write(uploaded_raw.getbuffer())

                                # 통합 로우데이터 대응: 강사명 필터링
                                filtered_temp_path = os.path.join(tmpdir, f"filtered_{i}{raw_ext}")
                                actual_raw_path = filter_raw_data_by_instructor(raw_temp_path, filtered_temp_path, inst_name)

                                inst_config = copy.deepcopy(config_dict)
                                inst_config["basic"]["강사명"] = inst_name

                                output_temp_path = os.path.join(tmpdir, f"output_{i}.pptx")

                                streamlit_log(f"[{inst_name}] 결과보고서 생성 중...")
                                generate_report(
                                    raw_path=actual_raw_path,
                                    config=inst_config,
                                    template_path=template_temp_path,
                                    output_path=output_temp_path,
                                    drop_no_opinion=not keep_no_opinion,
                                    log=streamlit_log
                                )

                                if os.path.exists(output_temp_path):
                                    with open(output_temp_path, "rb") as f:
                                        pptx_results[inst_name] = f.read()

                        # 엑셀 파일 업로드 모드
                        else:
                            uploaded_raw = list(valid_raw_files.values())[0]
                            raw_ext = os.path.splitext(uploaded_raw.name)[1]
                            raw_temp_path = os.path.join(tmpdir, f"raw_data{raw_ext}")

                            with open(raw_temp_path, "wb") as f:
                                f.write(uploaded_raw.getbuffer())

                            config_temp_path = os.path.join(tmpdir, f"config{os.path.splitext(config_file.name)[1]}")
                            with open(config_temp_path, "wb") as f:
                                f.write(config_file.getbuffer())

                            # 엑셀 모드에서도 강사명을 추출하여 통합 데이터 필터링 수행
                            inst_name = "강사"
                            try:
                                import openpyxl
                                wb = openpyxl.load_workbook(config_temp_path, data_only=True)
                                if "기본정보" in wb.sheetnames:
                                    ws = wb["기본정보"]
                                    for r in range(1, ws.max_row + 1):
                                        if ws.cell(row=r, column=1).value == "강사명":
                                            inst_name = str(ws.cell(row=r, column=2).value or "강사").strip()
                                            break
                                wb.close()
                            except Exception:
                                pass

                            filtered_temp_path = os.path.join(tmpdir, f"filtered_data{raw_ext}")
                            actual_raw_path = filter_raw_data_by_instructor(raw_temp_path, filtered_temp_path, inst_name)

                            output_temp_path = os.path.join(tmpdir, "output.pptx")

                            streamlit_log("[엑셀 설정] 결과보고서 생성 중...")
                            generate_report(
                                raw_path=actual_raw_path,
                                config=config_temp_path,
                                template_path=template_temp_path,
                                output_path=output_temp_path,
                                drop_no_opinion=not keep_no_opinion,
                                log=streamlit_log
                            )

                            if os.path.exists(output_temp_path):
                                with open(output_temp_path, "rb") as f:
                                    pptx_results[inst_name] = f.read()

                        if pptx_results:
                            st.success("🎉 결과보고서 생성이 완료되었습니다!")

                            # 강사별 다운로드 버튼을 렌더링
                            title_base = "결과보고서"
                            if config_dict and (config_dict["basic"].get("표지_제목2") or config_dict["basic"].get("표지_제목1")):
                                title_base = config_dict["basic"].get("표지_제목2") or config_dict["basic"].get("표지_제목1")
                            
                            # 특수문자 제거
                            title_base = re.sub(r'[\/:*?"<>| ]', '_', title_base)

                            with download_area.container():
                                st.markdown("#### 📥 생성된 결과보고서 다운로드")
                                cols = st.columns(len(pptx_results))
                                for idx, (name, pptx_bytes) in enumerate(pptx_results.items()):
                                    with cols[idx % len(cols)]:
                                        dl_filename = f"{title_base}_{name}_결과보고서.pptx"
                                        st.download_button(
                                            label=f"📥 [{name}] 다운로드 (.pptx)",
                                            data=pptx_bytes,
                                            file_name=dl_filename,
                                            mime="application/vnd.openxmlformats-officedocument.presentationml.presentation",
                                            use_container_width=True,
                                            key=f"dl_btn_{idx}"
                                        )
                        else:
                            st.error("❌ 결과 파일 생성에 실패했습니다. 로그를 확인하세요.")

                    except Exception as e:
                        st.error(f"❌ 생성 도중 오류가 발생했습니다: {str(e)}")
                        st.exception(e)